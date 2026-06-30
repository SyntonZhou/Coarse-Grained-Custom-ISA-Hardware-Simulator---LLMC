"""topology.py -- explicit DMA / NoC / HBM-PC topology model.

This file separates three concepts that are easy to conflate in the BD:

  * core/DMA initiators     : which DMA and NoC NMU a compute core uses;
  * NoC NMU ports           : HBM00_AXI_nmu, HBM01_AXI_nmu, ...;
  * physical HBM pseudo-PCs : HBM0_PC0 .. HBM7_PC1.

The default topology matches the current single-core BD/QoS interpretation:
HBM00/HBM01 are the LLMCORE bulk-data DMA read/write NMUs and can reach the
14 data PCs; HBM02/HBM03 are the TinyRISC-V paths and are restricted to the two
control/instruction PCs.  For a future 8-core design, set independent_dma=True
only if the RTL really instantiates independent DMA engines or equivalent
streaming movers per core.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Dict, Iterable, List, Tuple

from .memory import DEFAULT_REGIONS


PC_NAMES: List[str] = [name for _, name in DEFAULT_REGIONS]

# Physical interpretation used by the QoS table discussion.
PC_TO_PHYS: Dict[str, str] = {
    "HBM0_PC0": "hbm_st0.ch0", "HBM0_PC1": "hbm_st0.ch1",
    "HBM1_PC0": "hbm_st0.ch2", "HBM1_PC1": "hbm_st0.ch3",
    "HBM2_PC0": "hbm_st0.ch4", "HBM2_PC1": "hbm_st0.ch5",
    "HBM3_PC0": "hbm_st0.ch6", "HBM3_PC1": "hbm_st0.ch7",
    "HBM4_PC0": "hbm_st1.ch0", "HBM4_PC1": "hbm_st1.ch1",
    "HBM5_PC0": "hbm_st1.ch2", "HBM5_PC1": "hbm_st1.ch3",
    "HBM6_PC0": "hbm_st1.ch4", "HBM6_PC1": "hbm_st1.ch5",
    "HBM7_PC0": "hbm_st1.ch6", "HBM7_PC1": "hbm_st1.ch7",
}

CONTROL_PCS = ["HBM0_PC0", "HBM0_PC1"]
DATA_PCS = [pc for pc in PC_NAMES if pc not in CONTROL_PCS]


@dataclass(frozen=True)
class CorePort:
    """How one compute core reaches HBM through DMA/NoC."""
    core: int
    read_dma: str
    write_dma: str
    read_nmu: str
    write_nmu: str


@dataclass(frozen=True)
class NoCEdge:
    """One routable NMU -> physical PC path."""
    nmu: str
    pc: str
    latency_cycles: float
    read_qos_MBps: float = 100.0
    write_qos_MBps: float = 100.0
    # Runtime effective cap for this path.  QoS numbers are treated as weights,
    # not hard runtime limits.
    path_bw_GBps: float = 16.0

    @property
    def phys(self) -> str:
        return PC_TO_PHYS.get(self.pc, self.pc)


@dataclass
class HardwareTopology:
    core_ports: Dict[int, CorePort]
    edges: Dict[Tuple[str, str], NoCEdge]
    pc_names: List[str] = field(default_factory=lambda: list(PC_NAMES))
    independent_dma: bool = True
    name: str = "versal-hbm-qos-derived"

    def port_for_core(self, core: int) -> CorePort:
        if core in self.core_ports:
            return self.core_ports[core]
        # Safe fallback for analyses that schedule more cores than the topology
        # was explicitly created for.
        return CorePort(core, f"dma{core}_mm2s", f"dma{core}_s2mm",
                        "HBM00_AXI_nmu", "HBM01_AXI_nmu")

    def edge(self, nmu: str, pc: str) -> NoCEdge | None:
        return self.edges.get((nmu, pc))

    def reachable_pcs(self, nmu: str) -> List[str]:
        return [pc for (src, pc), _ in self.edges.items() if src == nmu]

    def check_reachable(self, nmu: str, pc: str) -> bool:
        return (nmu, pc) in self.edges

    def to_report(self) -> Dict[str, object]:
        return {
            "name": self.name,
            "independent_dma": self.independent_dma,
            "cores": {c: vars(p) for c, p in sorted(self.core_ports.items())},
            "reachable": {nmu: self.reachable_pcs(nmu)
                          for nmu in sorted({n for n, _ in self.edges})},
        }


def _latency_for_data_pc(pc: str) -> float:
    """QoS-table-inspired latency for HBM00/HBM01 to data PCs."""
    table = {
        "HBM5_PC0": 24.0, "HBM5_PC1": 24.0,  # st1.ch2/ch3
        "HBM6_PC0": 28.0, "HBM6_PC1": 28.0,  # st1.ch4/ch5
        "HBM4_PC0": 32.0, "HBM4_PC1": 32.0,  # st1.ch0/ch1
        "HBM3_PC0": 36.0, "HBM3_PC1": 36.0,  # st0.ch6/ch7
        "HBM7_PC0": 36.0, "HBM7_PC1": 36.0,  # st1.ch6/ch7
        "HBM2_PC0": 44.0, "HBM2_PC1": 44.0,  # st0.ch4/ch5
        "HBM1_PC0": 48.0, "HBM1_PC1": 48.0,  # st0.ch2/ch3
    }
    return table.get(pc, 48.0)


def default_versal_hbm_topology(n_cores: int = 1, *, independent_dma: bool = True) -> HardwareTopology:
    """Build a conservative topology from the current BD/QoS interpretation."""
    edges: Dict[Tuple[str, str], NoCEdge] = {}

    # LLMCORE bulk data paths: HBM00 = DMA MM2S read, HBM01 = DMA S2MM write.
    for nmu in ("HBM00_AXI_nmu", "HBM01_AXI_nmu"):
        for pc in DATA_PCS:
            lat = _latency_for_data_pc(pc)
            edges[(nmu, pc)] = NoCEdge(nmu, pc, latency_cycles=lat,
                                       read_qos_MBps=100.0, write_qos_MBps=100.0)

    # TinyRISC-V/control paths: restricted to two control/instruction PCs.
    for nmu in ("HBM02_AXI_nmu", "HBM03_AXI_nmu"):
        for pc in CONTROL_PCS:
            edges[(nmu, pc)] = NoCEdge(nmu, pc, latency_cycles=24.0,
                                       read_qos_MBps=200.0, write_qos_MBps=200.0)

    # CIPS/CPM/PCIe NoC ingress.  It can be treated as broad but high-latency.
    for nmu in ("S00_AXI_nmu", "S01_AXI_nmu"):
        for pc in PC_NAMES:
            edges[(nmu, pc)] = NoCEdge(nmu, pc, latency_cycles=160.0,
                                       read_qos_MBps=100.0, write_qos_MBps=100.0)

    core_ports: Dict[int, CorePort] = {}
    for c in range(n_cores):
        rd_dma = f"dma{c}_mm2s" if independent_dma else "dma0_mm2s"
        wr_dma = f"dma{c}_s2mm" if independent_dma else "dma0_s2mm"
        core_ports[c] = CorePort(c, rd_dma, wr_dma,
                                 read_nmu="HBM00_AXI_nmu",
                                 write_nmu="HBM01_AXI_nmu")

    return HardwareTopology(core_ports=core_ports, edges=edges,
                            independent_dma=independent_dma)


def pcs_for_nmu(topology: HardwareTopology, nmu: str) -> Iterable[str]:
    return topology.reachable_pcs(nmu)

# ---------------------------------------------------------------------------
# Optional QoS-table importer
# ---------------------------------------------------------------------------
def _nmu_short_name(full_name: str) -> str:
    return full_name.rstrip('/').split('/')[-1]


def _pc_from_qos_target(full_name: str) -> str | None:
    """Map a QoS target path to an HBMx_PCy name.

    Example target:
      axi_noc_0/inst/MC_hbmc/inst/hbm_st1/I_hbm_chnl2
    maps to:
      hbm_st1.ch2 -> HBM5_PC0
    """
    import re
    m = re.search(r"hbm_st([01]).*?hbm_chnl(\d+)", full_name)
    if not m:
        return None
    phys = f"hbm_st{int(m.group(1))}.ch{int(m.group(2))}"
    inv = {v: k for k, v in PC_TO_PHYS.items()}
    return inv.get(phys)


def topology_from_qos_xlsx(path: str, n_cores: int = 1, *,
                           independent_dma: bool = True,
                           fallback: bool = True) -> HardwareTopology:
    """Build HardwareTopology from Vivado NoC QoS XLSX export.

    The Vivado table is hierarchical: a row ending in ``*_nmu`` introduces the
    source NMU, and following rows list target NSU/HBM channels.  Duplicate
    edges can appear; this importer keeps the lower latency and higher QoS
    requirements for the same (NMU, PC) pair.
    """
    try:
        import openpyxl
    except Exception as e:  # pragma: no cover - optional dependency guard
        if fallback:
            return default_versal_hbm_topology(n_cores, independent_dma=independent_dma)
        raise RuntimeError("openpyxl is required to import QoS XLSX tables") from e

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    edges: Dict[Tuple[str, str], NoCEdge] = {}
    current_nmu: str | None = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name:
            continue
        name = str(name)
        short = _nmu_short_name(name)
        if short.endswith("_nmu"):
            current_nmu = short
            continue
        if current_nmu is None:
            continue
        pc = _pc_from_qos_target(name)
        if pc is None:
            continue
        try:
            rbw = float(row[2]) if row[2] not in (None, "") else 100.0
        except Exception:
            rbw = 100.0
        try:
            rlat = float(row[4]) if row[4] not in (None, "") else 0.0
        except Exception:
            rlat = 0.0
        try:
            wbw = float(row[6]) if row[6] not in (None, "") else rbw
        except Exception:
            wbw = rbw
        try:
            wlat = float(row[8]) if row[8] not in (None, "") else rlat
        except Exception:
            wlat = rlat
        lat = max(rlat, wlat)
        key = (current_nmu, pc)
        old = edges.get(key)
        if old is None:
            edges[key] = NoCEdge(current_nmu, pc, latency_cycles=lat,
                                 read_qos_MBps=rbw, write_qos_MBps=wbw)
        else:
            edges[key] = NoCEdge(current_nmu, pc,
                                 latency_cycles=min(old.latency_cycles, lat),
                                 read_qos_MBps=max(old.read_qos_MBps, rbw),
                                 write_qos_MBps=max(old.write_qos_MBps, wbw),
                                 path_bw_GBps=old.path_bw_GBps)

    if not edges:
        if fallback:
            return default_versal_hbm_topology(n_cores, independent_dma=independent_dma)
        raise RuntimeError(f"no NMU->HBM edges found in QoS table {path!r}")

    core_ports: Dict[int, CorePort] = {}
    for c in range(n_cores):
        rd_dma = f"dma{c}_mm2s" if independent_dma else "dma0_mm2s"
        wr_dma = f"dma{c}_s2mm" if independent_dma else "dma0_s2mm"
        core_ports[c] = CorePort(c, rd_dma, wr_dma,
                                 read_nmu="HBM00_AXI_nmu",
                                 write_nmu="HBM01_AXI_nmu")
    return HardwareTopology(core_ports=core_ports, edges=edges,
                            independent_dma=independent_dma,
                            name=f"qos-xlsx:{path}")
